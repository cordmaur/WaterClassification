# ---------------------------------------------------
# Functions to manipulate the Raw Radiometry Data (Ld, Lu, Ed)
# ---------------------------------------------------
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from io import StringIO
from functools import partial
import shutil
import pyodbc
import os
from WaterClassification.core import plot_figures, plot_reflectances2, create_interpolated_columns, normalize
import plotly.graph_objects as go
import plotly.io as pio
import plotly.express as px
import configparser

from openpyxl import load_workbook


# ---------------------------------------------------
# Basic Utility functions
# ---------------------------------------------------
def listify(*args):
    result = []
    for arg in args:
        result.append(arg if isinstance(arg, slice) or isinstance(arg, list) else [arg])
    if len(result) == 1:
        return result[0]
    else:
        return tuple(result)


def slice_df(df, subset, column_name='DateTime'):
    if subset is not None:
        subset = listify(subset)

        df = df.set_index(column_name, drop=True)

        if (column_name == 'DateTime') and isinstance(subset, list):
            subset = pd.DatetimeIndex(subset)

        df = df.loc[subset]
        df = df.reset_index().rename(columns={'index': column_name})

    return df


def intersection(*lsts):
    inter = [value for value in lsts[0] if value in lsts[1]]
    if len(lsts) > 2:
        for lst in lsts[2:]:
            inter = [value for value in inter if value in lst]

    return inter


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index=True, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_file_datetime(file: Path, ft='%Y-%m-%d %H:%M:%S') -> str:
    """Get the formatted datetime of a specified file"""

    # if there is a wildcard '*' in the name, search for the first occurence
    if '*' in file.stem:
        files = [f for f in file.parent.glob(file.name)]
        if len(files) > 0:
            file = files[0]
        else:
            return 'Not Found'

    if file.exists():
        dt = datetime.fromtimestamp(file.stat().st_mtime)
        return dt.strftime(ft)
    else:
        return 'Not Found'

# ---------------------------------------------------
# Basic File Manipulation functions
# ---------------------------------------------------
def match_files_names(files, names_lst):
    """
    Given two lists (files and names) check if all the names appear in the files.
    This function serve to check for the completeness of the directory.
    """
    stems = [f.stem for f in files]

    for name in names_lst:
        if name not in stems:
            return False
    return True


def get_complete_dirs(base_dir, names_lst):
    """
    Go through the subdirectories in base_dir and return those that are completed, that contains all the
    files in names_lst.
    """
    # get all the subdirectories in the base_dir
    dirs = [d for d in base_dir.rglob("*") if d.is_dir()]

    complete_dirs = []
    # look for directories that have the "full structure"
    for d in dirs:
        files = [f for f in d.iterdir() if not f.is_dir()]

        if match_files_names(files, names_lst):
            complete_dirs.append(d)

    return complete_dirs


def check_file(path, file_name):
    """Check if a file exists in the given path"""
    path = Path(path)

    if path.exists():
        for f in path.iterdir():
            if file_name in f.name:
                return f

    return False


def copy_dir(old_dir, new_dir, pattern='*'):
    old_dir = Path(old_dir)
    new_dir = Path(new_dir)
    new_dir.mkdir(parents=True, exist_ok=True)

    for old_f in old_dir.glob(pattern):
        if not old_f.is_dir():
            new_f = new_dir / old_f.name
            shutil.copy(old_f, new_f)


def get_file_by_suffix(path, stem, suffixes=None):
    """Get the file that matches the suffix in the order of preference of suffixes"""
    if suffixes is None:
        suffixes = ['.csv', '.txt', '.mlb']

    for suffix in suffixes:
        f = check_file(path, stem+suffix)
        if f:
            return f

    return False


def backup(fn):
    if fn.exists():
        if not fn.with_suffix('.bak').exists():
            shutil.copy(fn, fn.with_suffix('.bak'))
    else:
        print(f'File {fn} does not exists to be backed up')


def rename_files(base_dir, old_name, new_name, recursive=True, match_suffix=True):
    base_dir = Path(base_dir)

    # get all files
    if recursive:
        files = [f for f in base_dir.rglob('*') if not f.is_dir()]
    else:
        files = [f for f in base_dir.glob('*') if not f.is_dir]

    # filter files
    if match_suffix:
        files = [f for f in files if f.name == old_name]
    else:
        files = [f for f in files if f.stem == old_name]

    for file in files:
        if match_suffix:
            file.rename(file.with_name(new_name))
        else:
            file.rename(file.with_name(new_name + file.suffix))


# ---------------------------------------------------
# Date Time Conversion Functions
# ---------------------------------------------------
def from_excel_date(ordinal, epoch=datetime(1900, 1, 1)):
    # Adapted from above, thanks to @Martijn Pieters

    if ordinal > 59:
        ordinal -= 1  # Excel leap year bug, 1900 is not a leap year!
    in_days = int(ordinal)
    frac = ordinal - in_days
    in_secs = int(round(frac * 86400.0))

    return epoch + timedelta(days=in_days - 1, seconds=in_secs) # epoch is day 1


def to_excel_date(dt, epoch=datetime(1900, 1, 1)):
    td = dt - epoch
    return round(td.days + 2 + td.seconds/86400, 6)


# ---------------------------------------------------
# Radiometry Manipulation from TXT/MLB
# ---------------------------------------------------
def open_radiometry(file):
    file = Path(file)
    # print(f'Opening radiometry {file}')
    # Get the data frames from the file
    if file.suffix in ['.txt', '.bak']:
        metadata = pd.read_csv(file, header=None, sep='\t', nrows=18).drop(columns=1)
        rdmtry = pd.read_csv(file, sep='\t', skiprows=19, header=1, parse_dates=True, skip_blank_lines=True)

    elif file.suffix == '.mlb':
        metadata = pd.read_fwf(file, widths=[1, 19, 9, 50], header=None, nrows=18).drop(columns=[0, 2])
        metadata = metadata.rename(columns={1: 0, 3: 2})
        rdmtry = pd.read_fwf(file, skiprows=19, header=1, engine='python', skip_blank_lines=True)
        # return rdmtry, metadata
    else:
        # Suffix doesn't implemented
        print(f'Open radiometry function cannot open {file.suffix} suffix')
        return None, None

    # Adjust metadata (header of the file)
    for i in range(2, len(metadata.columns)):
        metadata[2] = metadata[2] + ' ' + metadata[i+1].replace(np.nan, '')
        metadata = metadata.drop(columns=i+1)

    metadata[2] = metadata[2].str.strip()
    metadata.rename(columns={0: 'key', 2: 'value'}, inplace=True)

    # Adjust the radiometry measurements
    rdmtry = rdmtry.drop(columns=['NaN.1', 'NaN.2', 'NaN.3']).rename(columns={'NaN': 'DateTime'})

    # Convert the DateTime from excel to understandable format
    rdmtry['DateTime'] = rdmtry['DateTime'].map(from_excel_date)
    rdmtry.sort_values(by='DateTime', inplace=True)

    return rdmtry, metadata


def get_date(d, dt_type='start', source='txt', base_name='Ed spectrum LO', **kwargs):
    if source == 'txt':
        f = get_file_by_suffix(d, base_name, suffixes=['.txt', '.mlb'])
        if f is not None:
            rdmtry, _ = open_radiometry(f)
        else:
            return None

        # get the first cell (first date) or the last if end date
        return rdmtry.iloc[0, 0] if dt_type == 'start' else rdmtry.iloc[-1, 0]
    else:
        print(f"Function get_date only supports .txt source, but {source} found")
        # path = Path(d)
        # dates = get_calibrated_dates(path, **kwargs)
        # if len(dates) > 0:
        #     return dates[0] if dt_type == 'start' else dates[-1]
        # else:
        #     return None


def get_number_measurements(d, rdmtry_type='Ed', base_name='spectrum LO', sep=' '):
    f = get_file_by_suffix(d, rdmtry_type + sep + base_name)
    rdmtry, _ = open_radiometry(f)

    return len(rdmtry)


def check_structure(base_dir, prefixes, sep='_', base_name='spectrum_LO', dir_level=0):
    """
    Create a data frame structure based on the sub directories and the TXT/MLB files
    The prefix are the Ed Ld and Lu to search for in the names of the files
    Base name is the any constant the operator has used to save the files
    Separator is the connecting character between the prefix and the base_name (ex. Ld_spectrum LO)
    The file used to grab the dates will be the one with the first prefix
    """

    # first, we will create a filename list
    names_lst = [prefix + sep + base_name for prefix in prefixes]

    complete_dirs = get_complete_dirs(base_dir=base_dir, names_lst=names_lst)

    df = pd.DataFrame(data=complete_dirs, columns=['Dir'])

    df['Group'] = None
    df['Area'] = df['Dir'].map(lambda x: x.parts[-3+dir_level])
    df['Station'] = df['Dir'].map(lambda x: x.parts[-2+dir_level])

    # Fill start and end datetime for the measurement
    df['Start_Date'] = df['Dir'].map(partial(get_date, base_name=names_lst[0], dt_type='start'))
    df['End_Date'] = df['Dir'].map(partial(get_date, base_name=names_lst[0], dt_type='end'))
    df['Duration'] = df['End_Date'] - df['Start_Date']

    # Get the number of measurements for each sensor (just for consistency checks)
    df['Ed_measurements'] = df['Dir'].map(partial(get_number_measurements, rdmtry_type='Ed',
                                                  sep=sep, base_name=base_name))
    df['Ld_measurements'] = df['Dir'].map(partial(get_number_measurements, rdmtry_type='Ld',
                                                  sep=sep, base_name=base_name))
    df['Lu_measurements'] = df['Dir'].map(partial(get_number_measurements, rdmtry_type='Lu',
                                                  sep=sep, base_name=base_name))

    df['Description'] = df['Dir'].map(lambda x: x.stem)

    return df.sort_values(by='Start_Date')


def get_radiances(folder,  prefixes, sep='_', base_name='spectrum_LO', suffixes=None):

    radiances = {}
    metadatas = {}

    for prefix in prefixes:
        # get the filename corresponded by each prefix
        _base_name = 'interpolated' if 'Rrs' in prefix else base_name
        fn = get_file_by_suffix(folder, stem=(prefix + sep + _base_name), suffixes=suffixes)

        if fn:
            if 'Rrs' in prefix:
                rdmtry = pd.read_csv(fn, sep=';')
                rdmtry['DateTime'] = pd.to_datetime(rdmtry['DateTime'])
                meta = None
            else:
                rdmtry, meta = open_radiometry(fn)

            radiances.update({prefix: rdmtry})
            metadatas.update({prefix: meta})
        else:
            print(f'{prefix} not found in {folder}')

    return radiances, metadatas


# ---------------------------------------------------
# Radiometry Manipulation from MDB
# ---------------------------------------------------
def open_connection(mdb):
    return pyodbc.connect(r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=" + str(mdb))

# GET TABLE NAMES
# conn = pyodbc.connect(r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=D:/modelo.mdb")

# crsr = conn.cursor()
# for table_info in crsr.tables(tableType='TABLE'):
#     print(table_info.table_name)


def exec_query(conn, qry):
    cur = conn.execute(qry)
    return [row[0] if len(row) == 1 else row for row in cur.fetchall()]


def get_column_values(conn, tbl, column):
    qry = f'SELECT DISTINCT {column} FROM {tbl}'
    return exec_query(conn, qry)


def get_info(mdb):
    conn = open_connection(mdb)

    d = {'FileName': mdb,
         'Records': exec_query(conn, 'SELECT count(*) FROM tblData')[0],
         'Devices': get_column_values(conn, 'tblData', 'IDDevice'),
         'DataTypes': get_column_values(conn, 'tblData', 'IDDataType'),
         'DTypesSub1': get_column_values(conn, 'tblData', 'IDDataTypeSub1'),
         'Comments': get_column_values(conn, 'tblData', 'Comment'),
         'Comments1': get_column_values(conn, 'tblData', 'CommentSub1'),
         'Comments2': get_column_values(conn, 'tblData', 'CommentSub2'),

         }

    conn.close()
    return d


def count_data(conn, column='Data', disp_name='Count', where_clause=None, group_by='DateTime'):
    qry = f'SELECT tblData.{group_by}, Count({column}) AS {disp_name} FROM tblData '
    if where_clause:
        qry += f"WHERE {where_clause} "

    qry += f'GROUP BY tblData.{group_by}'

    df = pd.read_sql(qry, conn)
    df.set_index(group_by, inplace=True)
    return df.convert_dtypes()


def get_summary(mdb, detailed=False):
    conn = open_connection(mdb)
    df = count_data(conn, column='IDData', disp_name='Records')
    df['Inclination'] = count_data(conn, column='Data', where_clause="IDDataType='Inclination'")
    df['Pressure'] = count_data(conn, column='Data', where_clause="IDDataType='Pressure'")
    df['Spectrum'] = count_data(conn, column='Data', where_clause="IDDataType='SPECTRUM'")
    df['Lu Raw'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='RAW' "
                                           "AND Comment LIKE '%Lu%deck%'")
    df['Lu Cal'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='CALIBRATED' "
                                           "AND Comment LIKE '%Lu%deck%'")
    df['Ld Raw'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='RAW' "
                                           "AND Comment LIKE '%Ld%deck%'")
    df['Ld Cal'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='CALIBRATED' "
                                           "AND Comment LIKE '%Ld%deck%'")
    df['Ed Raw'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='RAW' "
                                           "AND Comment LIKE '%Ed%deck%'")
    df['Ed Cal'] = count_data(conn, column='Data',
                              where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='CALIBRATED' "
                                           "AND Comment LIKE '%Ed%deck%'")
    df['Reflectance'] = count_data(conn, column='Data',
                                   where_clause="IDDataType='SPECTRUM' AND IDDataTypeSub1='CALCULATED'")
    df.fillna(value=0, inplace=True)

    conn.close()
    if detailed:
        return df
    else:
        return df.groupby(by=['Inclination', 'Pressure', 'Spectrum', 'Lu Raw', 'Lu Cal',
                              'Ld Raw', 'Ld Cal', 'Ed Raw', 'Ed Cal', 'Reflectance']).count()

    ##################################################################################################################


# GET MEASUREMENT FUNCTIONS
def create_query(select_columns, conditions):
    # create the base query
    qry = f"SELECT {select_columns} FROM tblData "

    # add the conditions
    if conditions is not None:
        for i, (column, value) in enumerate(conditions.items()):
            logic = 'WHERE' if i == 0 else 'AND'
            qry += f"{logic} {column} LIKE '{value}' "

    return qry


def get_rows(mdb, select_columns, conditions=None, index_col=None, output_format='list'):
    """
    Given a dictionary of conditions {column: value, column:value, column: value}
    Return selected columns of the rows that meet these specifications.
    Output_format can be 'list' or 'dataframe'
    """

    # if mdb is passed as a string (or Path), open connection and close at the end
    conn = open_connection(mdb) if isinstance(mdb, (str, Path)) else mdb

    qry = create_query(select_columns, conditions)

    if output_format == 'list':
        result = exec_query(conn, qry)
    else:
        result = pd.read_sql(sql=qry, con=conn, index_col=index_col)

    if isinstance(mdb, (str, Path)):
        conn.close()

    return result


def get_date_times(mdb, r_types_descr):
    """
    Given a dictionary of {r_type: {column: value, column:value}, r_type: {column: value, column:value}}
    Get the dates that meet these specifications.
    """

    # if mdb is passed as a string (or Path), open connection and close at the end
    conn = open_connection(mdb) if isinstance(mdb, (str, Path)) else mdb

    dates_idx = None
    for r_type, conditions in r_types_descr.items():
        rows = get_rows(conn, 'DateTime', conditions=conditions)

        idx = pd.DatetimeIndex(rows)

        # check for duplicates
        if idx.has_duplicates:
            print(f'Duplicated datetime found for {r_type}. Keeping first occurrences.')

        dates_idx = idx if dates_idx is None else dates_idx.intersection(idx)

    if isinstance(mdb, (str, Path)):
        conn.close()

    return dates_idx


# def get_calibrated_dates(mdb, with_reflectance=True, r_type_column='Comment',
#                          r_types=None, station=None):
#     """Get the date_time for the already calibrated measurements that contains Ld, Ed and Lu.
#     The search for the radiometries will be done in the r_type_column and look for the r_types strings
#     If with_reflectance equals False, get all measurements regardless the presence of calculated reflectance
#     If we want to filter by a specific station we can provide column and name to filter
#     station=['Comment', 'SHH3Y']
#     """
#
#     if r_types is None:
#         r_types = ['Ed', 'Ld', 'Lu']
#
#     conn = open_connection(mdb)
#
#     qry = "SELECT distinct tblData.DateTime FROM "
#
#     if not with_reflectance:
#         qry += "(tblData INNER JOIN tblData AS tblData_1 ON tblData.DateTime = tblData_1.DateTime) " \
#                "INNER JOIN tblData AS tblData_2 ON tblData_1.DateTime = tblData_2.DateTime "
#     else:
#         qry += "((tblData INNER JOIN tblData AS tblData_1 ON tblData.DateTime = tblData_1.DateTime) "
#         qry += "INNER JOIN tblData AS tblData_2 ON tblData_1.DateTime = tblData_2.DateTime) "
#         qry += "INNER JOIN tblData AS tblData_3 ON tblData_2.DateTime = tblData_3.DateTime "
#
#     # where clauses
#     qry += f"WHERE tblData.{r_type_column} LIKE '%{r_types[0]}%' " \
#            f"AND tblData_1.{r_type_column} LIKE '%{r_types[1]}%' " \
#            f"AND tblData_2.{r_type_column} LIKE '%{r_types[2]}%' " \
#            f"AND tblData.IDDataTypeSub1='CALIBRATED' " \
#            f"AND tblData_1.IDDataTypeSub1='CALIBRATED' " \
#            f"AND tblData_2.IDDataTypeSub1='CALIBRATED' "
#
#     if with_reflectance:
#         qry += " AND tblData_3.IDDataTypeSub1='CALCULATED' AND tblData_3.IDDataType='SPECTRUM' "
#
#     if station:
#         qry += f" AND tblData.{station[0]} = '{station[1]}'"
#
#     dates = exec_query(conn, qry)
#     conn.close()
#
#     dates.sort()
#
#     return dates


# method_type = {'Ed': 'SAMIP Control', 'Ld': 'SAM Control', 'Lu': 'SAM Control'}


# def get_measurements(mdb, r_type='Ld', raw=False):
#     conn = open_connection(mdb)
#
#     comment = f'%{r_type}%'
#     sub_type1 = 'RAW' if raw else 'CALIBRATED'
#     qry = f"SELECT * from tblData where Comment LIKE '{comment}' AND IDDataTypeSub1='{sub_type1}' " \
#           f"AND IDDataType='SPECTRUM' AND IDMethodType='{method_type[r_type]}'"
#
#     results = pd.read_sql(qry, conn)
#
#
#     conn.close()
#
#     return results



# def get_measurements2(mdb, r_type_column='Comment', r_type='Ld', method_type=None, raw=False, station=None):
#     conn = open_connection(mdb)
#
#     comment = f'%{r_type}%'
#     sub_type1 = 'RAW' if raw else 'CALIBRATED'
#     qry = f"SELECT * from tblData where {r_type_column} LIKE '{comment}' AND IDDataTypeSub1='{sub_type1}' " \
#           f"AND IDDataType='SPECTRUM'"
#
#     if method_type is not None:
#         qry += f" AND IDMethodType='{method_type}'"
#
#     if station:
#         qry += f" AND tblData.{station[0]} = '{station[1]}'"
#
#     results = pd.read_sql(qry, conn)
#     conn.close()
#
#     return results

# def get_calibrated_measurements(mdb, r_type_descr):
#     """Get the measurements only that have Ld, Lu e Ed and calibrated values"""
#
#     idx = get_date_times(mdb, r_type_descr)
#     df = pd.DataFrame(index=idx)
#
#     df = pd.read_sql(qry, conn).set_index('DateTime')
#
#     lds = get_measurements2(mdb, r_type_column=r_type_column, r_type=r_type, method_type=method_type,
#                             station=station).set_index('DateTime')
#
#     df[lds.columns] = lds[lds.columns]
#
#     return df


# CREATE DF AND CSV FUNCTIONS
def parse_attributes(attrs):
    attrs_dic = {}
    for attr in attrs.splitlines():
        key, value = attr.split('=')
        attrs_dic.update({key.strip(): value.strip()})
    return attrs_dic


def parse_row(row):
    """Expand a measurement row with the attributes in the attributes field and return the row
    as a series with all attributes"""

    # Parse the attributes of this measurement and append to the series
    attrs = parse_attributes(row['Attributes'])
    series = row.append(pd.Series(attrs))

    return series


def create_header_df(row):
    """Create a header as a DataFrame with the information from the passed row"""

    # First, expand the row, by parsing all the attributes
    series = parse_row(row)

    # Get just the attributes for the header
    header_attrs = ['IDDevice', 'IDDataType', 'IDDataTypeSub1', 'IDDataTypeSub2', 'IDDataTypeSub3', 'RecordType',
                    'IDMethodType', 'MethodName']
    header_attrs += ['IntegrationTime', 'Unit1', 'Unit2', 'Unit3', 'Unit4', 'IDDataBack', 'IDDataCal', 'IDBasisSpec',
                     'PathLength', 'CalFactor']

    header = series.loc[header_attrs]

    # create a DataFrame with the header data
    data = {'Attribute': header.index, 'sep': '=', 'Value': header.values}
    df = pd.DataFrame(data=data).set_index('Attribute')

    return df


def create_measurements_df(measures, additional_header=False):
    measures_df = None
    wls = None

    for idx, row in measures.iterrows():
        data = row['Data']
        iodata = StringIO(data)

        attrs = parse_attributes(row['Attributes'])

        df = pd.read_csv(iodata, sep=' ', header=None)
        df.drop(index=0, columns=[0, 3, 4], inplace=True)

        df['DateTime'] = row.name
        pivotted = df.pivot(index='DateTime', columns=1, values=2)

        # get the columns that represent the wavelenghts
        wls = pivotted.columns

        # Create new columns
        pivotted['PositionLatitude'] = 0.0
        pivotted['PositionLongitude'] = 0.0
        pivotted['IntegrationTime'] = attrs['IntegrationTime']
        pivotted['Comment'] = row['Comment']
        pivotted['IDData'] = row['IDData']

        # reorder columns
        pivotted = pivotted[
            ['PositionLatitude', 'PositionLongitude', 'IntegrationTime'] + list(wls.values) + ['Comment', 'IDData']]

        measures_df = pivotted if measures_df is None else pd.concat([measures_df, pivotted], axis=0)

    if additional_header:
        c_header = ['DateTime', 'PositionLatitude', 'PositionLongitude', 'IntegrationTime'] + \
                   [f'c{str(i).zfill(3)}' for i in range(1, len(wls) + 1)] + ['Comment', 'IDData']
        c_header = '\t'.join(c_header)
        return measures_df, c_header
    else:
        return measures_df


def create_radiometry_csv(mdb, description, times=None):
    """Create a csv file in the same format as the .MLB or the .TXT of the MSDA_XE software"""

    # get the measurements in the MDB matching r_type_descr
    measures = get_rows(mdb, '*', description, output_format='dataframe', index_col='DateTime')

    # if there are specified times, use them to slice the dataframe
    if times is not None:
        measures = measures[measures.index.isin(times)]

        # now check if the passed dates were found in the MDB
        if len(measures) != len(times):
            print(f'Warning: {len(times)} times informed, but {len(measures)} were found')

    # Create the header text, pass the first row for it
    header_txt = create_header_df(measures.iloc[0]).to_csv(sep='\t', line_terminator='\n', header=False)

    # Create the measurements text
    measures_df, c_header = create_measurements_df(measures, additional_header=True)

    # Before writing to txt, convert the dates to Excel format
    measures_df.index = measures_df.index.map(to_excel_date)

    measures_txt = measures_df.to_csv(sep='\t', line_terminator='\n')

    for title in ['DateTime', 'PositionLatitude', 'PositionLongitude', 'IntegrationTime', 'Comment', 'IDData']:
        measures_txt = measures_txt.replace(title, 'NaN')

    return header_txt + '\n' + c_header + '\n' + measures_txt

# def export_mdb_csv(mdb_path, out_dir, r_types, with_reflectance=False, times=None):
#     for r_type in r_types:
#         txt = create_radiometry_csv(mdb_path,
#                                     r_type=r_type,
#                                     with_reflectance=with_reflectance,
#                                     times=times)
#         out_path = mdb_path.with_name(out_dir)
#         out_path.mkdir(parents=True, exist_ok=True)
#         with open(out_path / f'{r_type} spectrum LO.txt', "w") as text_file:
#             text_file.write(txt)


def export_mdb_csv2(mdb_path, r_types_descr, out_dir=None, create_measurement_dir=True, times=None):
    # if an output directory is specified, append it to the current path
    if out_dir is not None:
        out_path = mdb_path.with_name(out_dir)
    else:
        out_path = mdb_path.parent

    # if create_measurement_dir is true, create the datetime dir for the specific measurement
    if create_measurement_dir:
        if times is not None:
            dt = times[0]
        else:
            dt = get_date_times(mdb_path, r_types_descr)

        out_name = str(dt).replace('-', '').replace(' ', '-').replace(':', '')[:13]
        out_path = out_path / out_name

    print(f'Saving output file to {out_path}')
    out_path.mkdir(parents=True, exist_ok=True)

    # Loop through the r_types
    for r_type, description in r_types_descr.items():
        txt = create_radiometry_csv(mdb_path,
                                    description=description,
                                    times=times)

        with open(out_path / f'{r_type}_spectrum_LO.txt', "w") as text_file:
            text_file.write(txt)

    return out_path


# MDBs BATCH PROCESSING
def loop_mdbs(start_path, recursive=True):
    # make sure we have a Path object
    start_path = Path(start_path)

    # use the correct function to grab the files considering recursiveness
    scan_func = Path.rglob if recursive else Path.glob

    # Loop through all MDBs including subdirectories
    for mdb in scan_func(start_path, '*.mdb'):
        yield mdb


def scan_mdbs(start_path, recursive=True):
    data = {i: get_info(mdb) for i, mdb in enumerate(loop_mdbs(start_path=start_path, recursive=recursive))}
    return pd.DataFrame(data).T


def create_dates_summary(dates, with_reflectance=False):
    if len(dates) > 0:
        s = f'Found {len(dates)} measurements from {str(dates[0])} to {str(dates[-1])}: ' + (
            'Only dates with reflectance' if with_reflectance else 'Regardless reflectance')
    else:
        s = 'No calibrated dates found'
    return s


def process_mdbs(start_path, with_reflectance=False, recursive=True):
    for mdb in loop_mdbs(start_path=start_path, recursive=recursive):
        print('-' * 10)
        print(f'Processing mdb: {mdb}')

        dates = get_calibrated_dates(mdb, with_reflectance=with_reflectance)
        print(create_dates_summary(dates))

        if len(dates) > 0:
            out_name = str(dates[0]).replace('-', '').replace(' ', '-').replace(':', '')[:13]
            out_dir = mdb.with_name(out_name)
            out_dir.mkdir(parents=True, exist_ok=True)

            print(f'Saving output file to {out_dir}')

            for r_type in ['Ed', 'Lu', 'Ld']:
                txt = create_radiometry_csv(mdb, r_type=r_type)

                with open(str(out_dir / f'{r_type} spectrum LO.txt'), "w") as text_file:
                    text_file.write(txt)

        print('')


# ---------------------------------------------------
# Reflectance functions
# ---------------------------------------------------
def create_reflectance(folder, ro=0.028):
    radiances, _ = get_radiances(folder, prefixes=['Ld', 'Lu', 'Ed'])

    lu = radiances['Lu'].set_index('DateTime')  # .drop(columns=['NaN.4', 'NaN.5', 'NaN.6'])
    ld = radiances['Ld'].set_index('DateTime')  # .drop(columns=['NaN.4', 'NaN.5', 'NaN.6'])
    ed = radiances['Ed'].set_index('DateTime')  # .drop(columns=['NaN.4', 'NaN.5', 'NaN.6'])

    # Will use undocumented _get_numeric_data() to grab just numeric columns
    interp_lu = create_interpolated_columns(lu, create_id=False)._get_numeric_data()
    interp_ld = create_interpolated_columns(ld, create_id=False)._get_numeric_data()
    interp_ed = create_interpolated_columns(ed, create_id=False)._get_numeric_data()

    r_rs = (interp_lu - ro * interp_ld) / interp_ed

    # before returning, delete empty rows
    return r_rs.dropna(axis=0, how='all')


def save_reflectance(folder, ro=0.028):
    r_rs = create_reflectance(folder, ro=ro)
    r_rs.to_csv(folder / 'Rrs interpolated.txt', sep='\t')


def save_reflectances(df, ro=0.028):
    for idx, row in df.iterrows():
        save_reflectance(row['Path'], ro=ro)
        # r_rs = create_reflectance(row['Path'], ro=ro)
        #
        # r_rs.to_csv(row['Path']/'Rrs interpolated.txt', sep='\t')


# ---------------------------------------------------
# GET-DB Structure Functions
# ---------------------------------------------------
def create_getdb_structure(base_dir):
    base_dir = Path(base_dir)
    df = None
    areas = [f for f in base_dir.iterdir() if (not f.stem.startswith('_')) and f.is_dir()]

    for area in areas:
        stations = [f for f in area.iterdir() if (not f.stem.startswith('_')) and f.is_dir()]

        for station in stations:
            measures = [f.stem for f in station.iterdir() if not f.stem.startswith('_')]
            measures_df = pd.DataFrame(measures).rename(columns={0: 'Measurement'})
            measures_df['Path'] = [f for f in station.iterdir()]
            measures_df['Station'] = station.stem
            measures_df['Area'] = area.stem

            df = measures_df if df is None else pd.concat([df, measures_df], axis=0)

    df['StartDate'] = np.nan
    # df['Path'].map(partial(get_date, dt_type='start', source='txt', base_name='Ed spectrum LO'))
    df['EndDate'] = np.nan
    # df['Path'].map(partial(get_date, dt_type='end', source='txt', base_name='Ed spectrum LO'))

    df['Duration'] = np.nan  # df['EndDate'] - df['StartDate']

    # Get the number of measurements for each sensor (just for consistency checks)
    df['EdMeasures'] = np.nan  # df['Path'].map(partial(get_number_measurements, rdmtry_type='Ed'))
    df['LdMeasures'] = np.nan  # df['Path'].map(partial(get_number_measurements, rdmtry_type='Ld'))
    df['LuMeasures'] = np.nan  # df['Path'].map(partial(get_number_measurements, rdmtry_type='Lu'))

    df = df[['Area', 'Station', 'Measurement', 'StartDate', 'EndDate', 'LdMeasures', 'LuMeasures',
             'EdMeasures', 'Path']]

    return df.reset_index(drop=True)


def create_getdb_structure2(base_dir):
    base_dir = Path(base_dir)
    df = None
    areas = [f for f in base_dir.iterdir() if (not f.stem.startswith('_')) and f.is_dir()]

    for area in areas:
        stations = [f for f in area.iterdir() if (not f.stem.startswith('_')) and f.is_dir()]

        for station in stations:
            measures = [f.stem for f in station.iterdir() if not f.stem.startswith('_')]
            measures_df = pd.DataFrame(measures).rename(columns={0: 'Measurement'})
            measures_df['Path'] = [f for f in station.iterdir()]
            measures_df['Station'] = station.stem
            measures_df['Area'] = area.stem

            df = measures_df if df is None else pd.concat([df, measures_df], axis=0)

    return df.reset_index(drop=True)[['Path', 'Area', 'Station', 'Measurement']]


def load_getdb(base_dir, attrs=None, radiometries=None):
    if attrs is None:
        attrs = ['Id', 'StationName', 'Measurement', 'SPM', 'Start_Date', 'End_Date', 'Duration']
        attrs += ['Point_Description', 'Project', 'latitude', 'longitude', ]
        attrs += ['gmt_datetime', 'Use_SummerTime', 'Is_SummerTime', 'GMT']
        attrs += ['Ed_device', 'Ld_device', 'Lu_device']
        attrs += ['Rrs_measurements', 'Ed_measurements', 'Ld_measurements', 'Lu_measurements']

    df = create_getdb_structure2(base_dir)

    def load_metadata(row):
        fn = row['Path'] / 'metadata.txt'
        if fn.exists():
            config = configparser.ConfigParser()
            config.read(fn)

            attrs_values = [config.get('Metadata', attr, fallback=None) for attr in attrs]

        else:
            attrs_values = [None] * len(attrs)

        rename_dict = {i: name for i, name in enumerate(attrs)}
        return pd.Series(attrs_values).rename(rename_dict)

    meta_df = df.apply(load_metadata, axis=1)

    df[attrs] = meta_df[attrs]

    if 'SPM' in attrs:
        df['SPM'] = df['SPM'].replace('', np.nan).astype('float')

    # if no radiometry has been requested, return just the basic database structure
    if radiometries is None:
        return df

    # otherwise, load the radiometries of interest
    else:
        rdmtry = load_mean_radiometries(df, r_types=radiometries, path_column='Path')
        return df, rdmtry


def get_measures(df, area, station):
    return df[(df['Area'] == area) & (df['Station'] == station)]


def update_metadata_from_df(df, attrs_columns, dir_column='Path', section='Metadata'):
    def update_basic_metadata(row):
        row = row.fillna('')
        if not pd.isnull(row[dir_column]):
            out_path = Path(row[dir_column])

            if out_path.exists():
                try:
                    fn = out_path / 'metadata.txt'
                    config = configparser.ConfigParser()
                    config.read(fn)

                    config[section].update({column: str(row[column]) for column in attrs_columns})

                    with open(fn, 'w') as configfile:
                        config.write(configfile)

                    return 'Ok'
                except Exception as err:
                    return err

            else:
                return f'Folder {out_path} does not exists'

    return df.apply(update_basic_metadata, axis=1)


# ---------------------------------------------------
# Create GRAPHS
# ---------------------------------------------------
def plot_radiometry(rdmtry, subset=None, color=True, **kwargs):
    rdmtry = slice_df(rdmtry, subset=subset, column_name='DateTime').copy()

    numeric_columns = rdmtry._get_numeric_data().columns

    if color:
        rdmtry['DateTime'] = pd.to_datetime(rdmtry['DateTime'])
        rdmtry['J_Dt'] = rdmtry['DateTime'].map(lambda x: x.to_julian_date())
        color_column = 'J_Dt'
    else:
        color_column = None

    fig = plot_reflectances2(rdmtry,
                             bands=numeric_columns,
                             color=color_column,
                             hover_vars=['DateTime'],
                             log_color=False,
                             colorbar=False,
                             **kwargs)

    return fig


# ---------------------------------------------------
# Radiometry Class
# ---------------------------------------------------
class Radiometry:

    labels = {'Ed': {'y_axis': "Irradiance (mW/(m^2))",
                     'title': 'Irradiance (Ed)'},
              'Ld': {'y_axis': 'Radiance (mW/(m^2 sr))',
                     'title': 'Radiance (Ld)'},
              'Lu': {'y_axis': 'Radiance (mW/(m^2 sr))',
                     'title': 'Radiance (Lu)'},
              'Rrs': {'y_axis': 'Reflectance (sr^-1)',
                      'title': 'Reflectance (Rrs)'}}

    def __init__(self, radiances, metadata, interp_radiances, folder=None):
        self.radiances = radiances
        self.metadata = metadata
        self.folder = folder
        self.interp_radiances = interp_radiances

        self._subset = None

    @staticmethod
    def restore_backup(folder, prefixes=None, sep=' ', base_name='spectrum LO'):
        prefixes = ['Rrs', 'Ed', 'Ld', 'Lu'] if prefixes is None else prefixes
        folder = Path(folder)
        for prefix in prefixes:
            _base_name = 'interpolated' if prefix == 'Rrs' else base_name
            bn = folder / (prefix + sep + _base_name + '.bak')
            shutil.copy(bn, bn.with_suffix('.txt'))

    def update_metadata(self, subset=True):
        if self.metadata['Metadata'] is None:
            config = configparser.ConfigParser()
            config.add_section('Metadata')
            # config.add_section('MeanReflectance')

            self.metadata.update({'Metadata': config})
        else:
            config = self.metadata['Metadata']

        config.set('Metadata', 'Ed_device', self.metadata['Ed'].fillna('').set_index('key').loc['IDDevice'].value)
        config.set('Metadata', 'Ld_device', self.metadata['Ld'].fillna('').set_index('key').loc['IDDevice'].value)
        config.set('Metadata', 'Lu_device', self.metadata['Lu'].fillna('').set_index('key').loc['IDDevice'].value)

        config.set('Metadata', 'Rrs_Measurements', str(len(self['Rrs'])))
        config.set('Metadata', 'Ed_Measurements', str(len(self['Ed'])))
        config.set('Metadata', 'Ld_Measurements', str(len(self['Ld'])))
        config.set('Metadata', 'Lu_Measurements', str(len(self['Lu'])))

        dates = self.dates_range(string=False)
        config.set('Metadata', 'Start_Date', str(dates[0]))
        config.set('Metadata', 'End_Date', str(dates[1]))
        config.set('Metadata', 'Duration', str(dates[1] - dates[0]))

        # if subset:
        #     r_rs = self.get_radiometry_subset('Rrs', interpolated=True).mean()
        # else:
        #     r_rs = self.get_radiometry('Rrs', interpolated=True).mean()
        #
        # config['MeanReflectance'] = r_rs.to_dict()

        with open(self.folder / 'metadata.txt', 'w') as configfile:
            config.write(configfile)

    @staticmethod
    def open_metadata(folder):
        file = Path(folder)/'metadata.txt'
        if not file.exists():
            # print(f'Metadata file not found in {str(folder)}')
            return None
        else:
            config = configparser.ConfigParser()
            config.read(file)
            return config

    @classmethod
    def from_folder(cls, folder, prefixes=None, sep='_', base_name='spectrum_LO', read_backup=False,
                    load_interpolated=False, base_interpolated_name='_interpolated.csv'):
        """Open the radiometry on a specific folder. If prefixes are not specified, open the basic [Lu, Ld, Ed, Rrs]"""
        prefixes = ['Rrs', 'Ed', 'Lu', 'Ld', ] if prefixes is None else prefixes

        folder = Path(folder)

        suffixes = ['.bak', '.txt', '.mlb'] if read_backup else None

        radiances, metadatas = get_radiances(folder, prefixes=prefixes, sep=sep,
                                             base_name=base_name, suffixes=suffixes)

        interp_radiances = {}

        if load_interpolated:
            for prefix in prefixes:
                if prefix == 'Rrs':
                    continue

                fn = folder/(prefix + base_interpolated_name)
                rdmtry = pd.read_csv(fn, sep=';')
                rdmtry['DateTime'] = pd.to_datetime(rdmtry['DateTime'])
                interp_radiances.update({prefix: rdmtry})

        metadatas.update({'Metadata': cls.open_metadata(folder)})

        rdmtry = cls(radiances, metadatas, interp_radiances, folder)
        return rdmtry

    @classmethod
    def from_df_entry(cls, df, idx=None, area=None, station=None, measurement=None, prefixes=None, sep=' ',
                      base_name='spectrum_LO', base_interpolated_name='_interpolated.csv', read_backup=False):
        """Open a radiometry given an entry and a structured DataFrame.
        If idx is passed, it has precedence in selecting the item."""
        if idx is not None:
            row = df.loc[idx]
        else:
            row = df[(df['Area'] == area) & (df['Station'] == station) & (df['Measurement'] == measurement)].iloc[0]

        return cls.from_folder(row['Path'], prefixes=prefixes, sep=sep, base_name=base_name,
                               read_backup=read_backup, base_interpolated_name=base_interpolated_name)

    @classmethod
    def from_description(cls, db_path, area, station, measurement, prefixes=None, sep='_', base_name='spectrum_LO',
                         read_backup=False, load_interpolated=False, base_interpolated_name='_interpolated.csv'):
        bd_path = Path(db_path)
        folder = bd_path/area/station/measurement
        return cls.from_folder(folder, prefixes=prefixes, sep=sep, base_name=base_name, read_backup=read_backup,
                               load_interpolated=load_interpolated, base_interpolated_name='_interpolated.csv')

    @property
    def subset(self): return self._subset
    @subset.setter
    def subset(self, value):
        # if the value is a list of dates, check if they are present in all radiances and use
        # the intersection
        if isinstance(value, list):
            times = [value] + [df['DateTime'].astype('str').to_list() for df in self.radiances.values()]
            self._subset = intersection(*times)
        else:
            self._subset = value

    @property
    def station_name(self):
        area, local, measure = self.get_area_location_measurement()
        return area + '-' + local

    @property
    def spm(self):
        if self.metadata['Metadata'] is None:
            return None

        spm = self.metadata['Metadata'].get('Metadata', 'SPM', fallback=None)
        return None if (spm is None) or (spm == '') else float(spm)

    def get_wl_range(self, r_type):
        if r_type not in self.radiances.keys():
            print(f'{r_type} not in the radiances')
            return None
        else:
            return self[r_type].mean().dropna().index.min(), self[r_type].mean().dropna().index.max()

    def get_title_summary(self, subset):
        if subset and self.subset:
            if isinstance(self.subset, list):
                s = f'{len(self.subset)} measures from subset: {subset[0]}-{subset[-1]}'
            else:
                s = f'Subset: {self.subset}'
        else:
            s = f'{len(self.dates())} measures from {self.dates_range()}'

        s = f'SPM={self.spm} mg/l   :  ' + s
        return s

    def get_area_location_measurement(self):
        measurement = self.folder.stem
        location = self.folder.parent.stem
        area = self.folder.parent.parent.stem
        return area, location, measurement

    def add_mean_trace(self, fig, r_type, subset=True, color='Red'):
        rdmtry = self.get_radiometry(r_type, subset=subset) if subset else self[r_type]

        mean = rdmtry.mean()
        fig.add_trace(go.Scatter(x=mean.index,
                                 y=mean,
                                 name='Mean',
                                 line=dict(width=2.5, color=color)))

        std = rdmtry.std()
        fig.add_trace(go.Scatter(x=std.index,
                                 y=std+mean,
                                 name='+1 Std',
                                 line=dict(width=2.5, color='Blue', dash='dash')))

        fig.add_trace(go.Scatter(x=std.index,
                                 y=mean-std,
                                 name='-1 Std',
                                 line=dict(width=2.5, color='Blue', dash='dot')))
        return None

    def plot_radiometry(self, r_type='Rrs', subset=True, mean=False, **kwargs):
        subset = self.subset if subset else None

        fig = plot_radiometry(self[r_type], subset=subset, color=not mean, **kwargs)

        if mean:
            self.add_mean_trace(fig, r_type, subset, 'Red')

        title = self.labels[r_type]['title'] + '<br>'
        title += self.get_title_summary(subset=subset)

        fig.update_layout(
            showlegend=True,
            title=title,
            xaxis_title="Wavelength (nm)",
            yaxis_title= self.labels[r_type]['y_axis'],
            font=dict(
                family="Courier New, monospace",
                size=12,
                color="Black")
        )
        fig.update_xaxes(range=self.get_wl_range(r_type=r_type))

        return fig

    def plot_radiometries(self, subset=True, mean=False, **kwargs):
        figs = []

        subset = self.subset if subset else None

        for r_type in ['Rrs', 'Ed', 'Lu', 'Ld']:
            sub_fig = self.plot_radiometry(r_type=r_type, subset=subset, mean=mean, **kwargs)
            figs.append(sub_fig)

        fig = plot_figures(2, 2, figs,
                           titles=['Reflectance', 'Irradiance (Ed)', 'Radiance (Lu)', 'Radiance (Ld)'])

        # Adjust the title
        title = ' '.join(self.get_area_location_measurement())
        title += "<span style='font-size: 12px;'>"
        title += f'   ({self.folder})<br>'
        title += self.get_title_summary(subset=subset)
        title += '</span>'
        fig.update_layout(title=title)
        fig.update_xaxes(range=self.get_wl_range(r_type='Rrs'))

        return fig

    def save_radiometries_graph(self, folder, subset=True, mean=False, **kwargs):
        folder = Path(folder) if folder is not None else self.folder
        fig = self.plot_radiometries(subset=subset, mean=mean, **kwargs)
        area, location, measurement = self.get_area_location_measurement()
        fn = '_'.join([area, location, measurement]) + '.png'

        pio.write_image(fig, str(folder/fn), width=2000, height=1200, validate=False, engine='kaleido')

    def create_measurement_name(self):
        return self.dates_range()[0].replace('-', '').replace(' ', '-').replace(':', '')[:13]

    def dates(self):
        return self.radiances[next(iter(self.radiances.keys()))]['DateTime']

    def adjust_time(self, t_delta):
        """Adjust the datetime of the radiances by a timedelta. It is necessary to recreate reflectance afterwards"""
        for rd in self.radiances.values():
            rd['DateTime'] = rd['DateTime'] + t_delta

        new_folder = self.folder.with_name(self.create_measurement_name())
        self.folder.rename(new_folder)
        self.folder = new_folder

    def dates_range(self, string=True):
        dates = self.dates()
        if string:
            return str(dates.min()), str(dates.max())
        else:
            return dates.min(), dates.max()

    def get_sliced_radiance(self, date_range, r_type='Ed'):
        df = self.radiances[r_type].set_index('DateTime', drop=True)
        df = df.loc[date_range]

        return df.reset_index()

    def get_radiometry(self, r_type, subset=False, interpolated=False):
        if interpolated and (len(self.interp_radiances) == 0):
            self.create_interp_radiances()

        rep = self.radiances if (not interpolated) or (r_type == 'Rrs') else self.interp_radiances
        if r_type in list(rep.keys()):
            rdmtry = rep[r_type]

            if subset:
                return slice_df(rdmtry, self.subset, column_name='DateTime')
            else:
                return rdmtry

        else:
            print(f"No radiance '{r_type}'{':interpolated' if interpolated else ''} found")
            return None

    def apply_filter(self, date_range, min_value=-np.inf, max_value=np.inf, r_type='Ed', accum_filter=False):
        """The filter overrides the given _subset and establishes a new one.
        If date_range is None, keep all times in the list"""
        df = self.get_sliced_radiance(date_range, r_type=r_type) if date_range else self[r_type]

        subset = df[(df.max(axis=1) > min_value) & (df.max(axis=1) < max_value)]['DateTime'].astype('str').to_list()

        if accum_filter and self._subset:
            self.subset = intersection(subset, self._subset)
        else:
            self.subset = subset

        # subset = intersection(subset, self.radiances['Rrs']['DateTime'].astype('str').to_list())

        # if accum_filter and len(self._subset) > 0:

        # self._subset = intersection(subset, self._subset) if not accum_filter else intersection

    def _save_radiance(self, r_type, subset=True, save_backup=True):
        fn = self.folder / (r_type + '_spectrum_LO.txt')
        meta = self.metadata[r_type].copy()

        measures = self.get_radiometry(r_type, subset=subset,
                                       interpolated=False).set_index('DateTime', drop=True).copy()

        meta.insert(1, 'Symbol', '=')
        header_txt = meta.to_csv(sep='\t', line_terminator='\n', header=False, index=False)

        # Before writing to txt, convert the dates to Excel format
        measures.index = measures.index.map(to_excel_date)
        measures.insert(0, 'IntegrationTime', int(meta[meta['key']=='IntegrationTime']['value'].values))
        measures.insert(0, 'PositionLongitude', 0)
        measures.insert(0, 'PositionLatitude', 0)

        measures_txt = measures.to_csv(sep='\t', line_terminator='\n', na_rep='NaN')

        for title in ['DateTime', 'PositionLatitude', 'PositionLongitude', 'IntegrationTime', 'Comment', 'IDData']:
            measures_txt = measures_txt.replace(title, 'NaN')

        c_header = [f'c{str(i).zfill(3)}' for i in range(1, len(measures.columns) - 3 + 1)]
        c_header = ['DateTime', 'PositionLatitude', 'PositionLongitude', 'IntegrationTime'] + c_header
        c_header = c_header + ['Comment', 'IDData']

        c_header = '\t'.join(c_header)

        txt = header_txt + '\n' + c_header + '\n' + measures_txt

        if save_backup:
            backup(fn)

        with open(fn, "w") as text_file:
            text_file.write(txt)

    def _save_interpolated_radiance(self, r_type, subset=True, save_backup=True, interpolated_name='_interpolated.csv'):
        r_rs = self.get_radiometry(r_type, subset=subset, interpolated=True).set_index('DateTime', drop=True).copy()

        # create the filename
        fn = self.folder/(r_type + interpolated_name)

        # backup
        if save_backup:
            backup(fn)

        # save to csv
        r_rs.to_csv(fn, sep=';')

    def save_mean_radiometry(self, subset=True):
        """This function will create a MEAN spectrum LO.txt, with all the radiances available"""
        data = {}
        for r_type in list(self.radiances.keys()):

            rdmtry = self.get_radiometry(r_type, interpolated=True)
            rdmtry = rdmtry.median(skipna=False) if len(rdmtry) > 10 else rdmtry.mean()
            data.update({r_type: rdmtry})

        df = pd.DataFrame(data, index=data['Rrs'].index.rename('Radiometry')).T.dropna(how='all', axis=1)
        df.insert(0, 'DateTime', str(self.dates().mean())[:19])

        # create the filename
        fn = self.folder/'Mean_interpolated.csv'

        # backup
        backup(fn)

        df.to_csv(fn, sep=';')

        return df

    def save_radiometry(self, r_type, subset=True, save_interpolated=True, save_backup=True):

        if r_type == 'Rrs':
            self._save_interpolated_radiance(r_type, subset=subset, save_backup=save_backup)
        else:
            self._save_radiance(r_type, subset=subset, save_backup=save_backup)

            if save_interpolated:
                self._save_interpolated_radiance(r_type, subset=subset, save_backup=save_backup)

    def save_radiometries(self, r_types=None, subset=True, save_interpolated=True, save_backup=True):
        r_types = r_types if r_types is not None else self.radiances.keys()

        for r_type in r_types:
            self.save_radiometry(r_type, subset=subset, save_interpolated=save_interpolated, save_backup=save_backup)

        self.update_metadata(subset=subset)

    def create_interp_radiances(self, r_types=None, step=1, min_wl=None, max_wl=None):
        r_types = ['Ed', 'Lu', 'Ld'] if r_types is None else r_types
        for r_type in r_types:
            rd = self[r_type].set_index('DateTime')
            interp_rd = create_interpolated_columns(rd, create_id=False, step=step, min_col=min_wl, max_col=max_wl)\
                ._get_numeric_data()
            interp_rd.columns = interp_rd.columns.map(lambda x: str(x))
            self.interp_radiances.update({r_type: interp_rd.reset_index(drop=False)})

    def create_reflectance(self, ro=0.028, step=1, min_wl=None, max_wl=None):
        self.create_interp_radiances(step=step, min_wl=min_wl, max_wl=max_wl)

        ed = self.interp_radiances['Ed'].set_index('DateTime', drop=True)
        ld = self.interp_radiances['Ld'].set_index('DateTime', drop=True)
        lu = self.interp_radiances['Lu'].set_index('DateTime', drop=True)

        r_rs = (lu - ro * ld) / ed

        r_rs.dropna(axis=0, how='all', inplace=True)
        self.radiances.update({'Rrs': r_rs.reset_index(drop=False)})# before returning, delete empty rows

    def __getitem__(self, r_type):
        return self.get_radiometry(r_type, interpolated=False)

    def __repr__(self):
        s = f'Class Radiometry with {len(self.radiances)} radiances: {list(self.radiances.keys())} \n'
        s += f'Folder: {self.folder} \n'
        s += f'Date Range: {self.dates_range()} \n'
        s += f'Subset: '
        s += f'{len(self.subset)} items' if isinstance(self.subset, list) else f'{self.subset}'
        return s


class RadiometryDB:
    """Class responsible for loading the Radiometry Database and manipulating it."""

    config = {
        "control_file": 'BD-GET-Summary.xlsx',
        "mean_interpolated": 'Mean_interpolated.csv',
        "trios_output": 'Trios_Output.csv',
        "summary": '_summary.csv',
        "interpolated_name": '_interpolated.csv'
    }

    workbooks = {
        "data": {'name': 'GETDB-Summary', 'index': 'Id'},
        "stations": {'name': 'Stations', 'index': 'StationId'},
        "summertime": {'name': 'SummerTime', 'index': 'Id'},
        "granulometry": {'name': 'Granulometry', 'index': 'Id'},
        "kd": {'name': 'Kd', 'index': 'Id'}
    }

    all_radiometries = ['Rrs', 'Ed', 'Ld', 'Lu']

    default_attrs = ['Area', 'Station', 'Measurement', 'Start_Date', 'SPM', 'Status', 'Description']

    def __init__(self, path, config=None):
        """Open the radiometry DB located in specified folder"""
        self.path = Path(path)

        # update the main configuration of the class with the passed parameters
        if config is not None:
            self.config.update(config)

        # load main spreadsheets. They will be stored as instance attributes.
        # The attributes names and workbooks correspondence are defined in the workbooks dictionary
        self.load_control_file()

        # Join main data with the stations
        self.data = self.data.join(self.stations.set_index('Station_Name'), on='Station_Name')

        self.rdmtries = None

    # Loading functions
    def load_control_file(self):
        """Load the control file  and return data, stations and summertime dataframes.
        The configuration parameters shall be passed as a dict-like"""
        if check_file(self.path, self.config['control_file']):
            # Load all the workbooks defined in the class dictionary and add them as attributes to the class
            for workbook in self.workbooks:
                wb = pd.read_excel(self.path / self.config['control_file'],
                                   RadiometryDB.workbooks[workbook]['name'], engine='openpyxl')
                self.__setattr__(workbook, wb.set_index(RadiometryDB.workbooks[workbook]['index']))

            # data = pd.read_excel(path / config['control_file'], config['main_workbook'], engine='openpyxl')\
            #     .set_index('Id')
            # stations = pd.read_excel(path / config['control_file'], config['stations_workbook'], engine='openpyxl')\
            #     .set_index('StationId')
            # summertime = pd.read_excel(path / config['control_file'], config['summertime_workbook'], engine='openpyxl')
        else:
            print(f"Not possible to open control file {self.config['control_file']} at {self.path}")

    @staticmethod
    def _load_mean_radiometries(df, db_anchor, r_types='Rrs', path_column='Relative_Path', attrs=None,
                                mean_fname='Mean_interpolated.csv'):
        attrs = RadiometryDB.default_attrs if attrs is None else attrs

        db_anchor = Path(db_anchor)

        # convert the r_types into a list to be passed to the mapping function
        r_types = listify(r_types)

        def map_radiometry(row, rtypes, _attrs):
            rdmtry = pd.read_csv(db_anchor / row[path_column] / mean_fname, sep=';', index_col=0)

            # check if the r_type exists in the radiometry mean file
            for r_type in r_types:
                if not (r_type in rdmtry.index):
                    print(f'{r_type} not found in {row[path_column]}')

                if rdmtry.index.to_list().count(r_type) > 1:
                    print(f'More than one [{r_type}] in {row[path_column]}')

            return [row[_attrs].append(rdmtry.loc[r_type]) if r_type in rdmtry.index else None
                    for r_type in rtypes]

        idx_name = df.index.name
        df = df.reset_index(drop=False, inplace=False)
        result_df = df.apply(partial(map_radiometry, rtypes=r_types, _attrs=[idx_name]+attrs), axis=1)

        # as the result comes in list, we have to split it first
        result_df = pd.DataFrame(result_df.tolist())

        # create a dictionary with a dataframe for each interpolated radiometry
        result_dic = {r_type: pd.DataFrame(result_df[i].tolist()).dropna(axis=1, how='all').drop(columns='DateTime')
                      for i, r_type in enumerate(r_types)}

        # adjust index
        for r_type in r_types:
            result_dic[r_type].set_index('Id', inplace=True)

        return result_dic

    def load_radiometry(self, id=None, folder=None, area=None, station=None, measurement=None,
                        load_interpolated=True, read_backup=False):
        """Load a specific radiometry from the database and returns a Radiometry class
        The load radiometry function can work with either:
        - database Id
        - description: area, station and measurement ( in the format YYYYMMDD-hhmm)
        - folder: if we know the folder (absolute or relative) of the measurement in the database
        """
        if id is not None:
            path = self.path/self.data.loc[id]['Relative_Path']

        elif folder is not None:
            folder = Path(folder)
            path = folder if folder.is_absolute() else self.path/folder

        elif (area is not None) and (station is not None) and (measurement is not None):
            return Radiometry.from_description(self.path, area, station, measurement,
                                               load_interpolated=load_interpolated,
                                               read_backup=read_backup)

        else:
            print('Not enough info to load the radiometry. Check input arguments.')

        return Radiometry.from_folder(path,
                                      load_interpolated=load_interpolated,
                                      read_backup=read_backup)

    def load_radiometries(self, r_types=None, attrs=None, norm=False, funcs=None, qry=None):
        """
        Load the mean radiometries from all points in the database. It uses the summary csv's that have been
        previously created using create_summary_radiometries. The results are written as DataFrames into a dictionary.
        :param r_types: The radiometries to be loaded (ex. 'Rrs', 'Ed', 'Ld', 'Lu')
        :param attrs: Additional attributes to be loaded in the dataframe. Defaults to:
        ['Area', 'Station', 'Lat', 'Long', 'GMT', 'SummerTime', 'OBS', 'Relative_Path', 'Start_Date', 'End_Date',
        'Project', 'Description']
        :param funcs: List of functions to be applied to the resulting DataFrames
        :param norm: Include normalized bands with prefix 'n'
        :param qry: Additional query to select only desired status
        :return: A dictionary of radiometries. Each entry is a Dataframe with corresponding measurements.
        """
        attrs = ['Status', 'SPM', 'Area', 'Station', 'Lat', 'Long', 'GMT', 'SummerTime', 'OBS', 'Relative_Path',
                 'Start_Date', 'End_Date', 'Project', 'Description'] if attrs is None else listify(attrs)

        r_types = RadiometryDB.all_radiometries if r_types is None else listify(r_types)

        self.rdmtries = {r_type: pd.read_csv(self.path/f"{r_type}{self.config['summary']}", index_col='Id')
                         for r_type in r_types}

        for r_type in r_types:
            # make a join to add the attributes
            self.rdmtries[r_type] = self.rdmtries[r_type].join(self.data[attrs], lsuffix='_2')

            # create the normalized bands (with suffix)
            if norm:
                df_norm = normalize(self.rdmtries[r_type])
                bands = df_norm.columns[df_norm.columns.str.isdigit()]
                df_norm.rename(columns={band: f'n{band}' for band in bands}, inplace=True)

                self.rdmtries[r_type] = pd.concat([self.rdmtries[r_type], df_norm[[f'n{b}' for b in bands]]], axis=1)

            # apply post-processing functions
            if funcs:
                funcs = listify(funcs)
                for func in funcs:
                    func(self.rdmtries[r_type])

            if qry:
                self.rdmtries[r_type].query(qry, inplace=True)

        print(f'Radiometries {r_types} loaded in dictionary .rdmtries')

    # Properties
    @property
    def control_file(self): return self.path/self.config['control_file']

    def status(self):
        """Output the overall status of the database (number of records, update dates, etc.) """
        print(10*'-' + ' DataBase Status ' + 10*'-')
        print(f'Base Folder: {self.path}')
        print(f'Control File: {self.control_file.name} (Modified: {get_file_datetime(self.control_file)})')

        print('\n')
        print('Summary files:')

        # Get the summary files and plot their modified dates
        files = [f for f in self.path.glob('*' + self.config['summary'])]
        for file in files:
            print(f'{file.name}: Modified on {get_file_datetime(file)}')

        # Specifies the files to be included in the status
        files = {'Reflectance': 'Rrs_interpolated.csv',
                 'Mean Reflectance': 'Mean_interpolated.csv',
                 'OSOAA Reflectance': 'Rrs_OSOAA_interpolated.csv',
                 'Metadata': 'metadata.txt',
                 'Graph': '*.png'
                 }

        print('\n')
        print('Integrity/date table:')

        # Create a dictionary to receive the results from each check
        series = {}

        # Create a series with the absolute directories to iter
        paths = self.data['Relative_Path'].map(lambda x: self.path/x)

        # For each file to be checked, call the get_file_datetime function
        for name, file in files.items():
            series[name] = paths.map(lambda x: get_file_datetime(x/ file, ft='%Y-%m-%d')).value_counts()

        # Return a pandas dataframe with the results
        return pd.DataFrame(series).sort_index(ascending=False)

    def summary(self, by='Area', plot=False, bars=None, **kwargs):
        """Create a summary DataFrame to present all data stored in the database
        :param bars: The name of the bars to include in the plot.
        This can be 'SPM', 'POC', 'DOC', 'Chl-a', 'Granul', 'Kd'
        :param plot: Instead of returning a DataFrame, return a plotly bar fig
        """
        bars = ['SPM', 'POC', 'DOC', 'Chl-a', 'Granul', 'Kd'] if bars is None else bars

        # make a joined summary DataFrame
        summary = self.data.copy()

        summary = summary.join(self.granulometry['D10']).join(self.kd[650])
        summary.rename(columns={'D10': 'Granul', 650: 'Kd'}, inplace=True)

        summary = summary.groupby(by=by).count()[bars]

        summary.sort_values(by=bars, ascending=False, inplace=True)
        if plot:
            return px.bar(summary, y=bars, barmode='group', **kwargs)
        else:
            return summary

    # Internal Functions
    def get_measurement_id(self, station_name, measurement):
        measurement = self.data[(self.data['Station_Name'] == station_name) &
                                (self.data['Measurement'] == measurement)]

        return int(measurement.index.values) if len(measurement) > 0 else None

    def get_station_id(self, station_name):
        station = self.stations[self.stations['Station_Name'] == station_name]
        return int(station.index.values) if len(station) > 0 else None

    def insert_station(self, row, overwrite=False):
        idx = self.get_station_id(row['Station_Name'])

        if idx is not None:
            if overwrite:
                print(f"*** Overwriting existing station {row['Station_Name']}(Id: {idx})")
                self.stations.loc[idx] = row
            else:
                print(f"*** Station {row['Station_Name']}(Id: {idx}) already exists. Skipping it.")
        else:
            row.name = self.stations.index.max() + 1
            print(f'Appending new station - Id={row.name}')
            self.stations = self.stations.append(row)

    def insert_measurement(self, row, overwrite=False):
        idx = self.get_measurement_id(row['Station_Name'], row['Measurement'])

        if idx is not None:
            if overwrite:
                print(f'*** Overwriting existing record. Id={idx}')
                self.data.loc[idx] = row
            else:
                print(f"*** Record Id={idx} already exists. Skipping it.")

        else:
            row.name = self.data.index.max() + 1
            print(f'Appending new data. Id={row.name}')
            self.data = self.data.append(row)

        return idx

    # Processing functions
    def loop_database(self, funcs, qry_str=None):
        """Loop the database and applies several callback functions to each row.
        The subset can be specified by a query string like: Id > 400 or SPM < 8.0 for example."""
        # if filters is not None:
        #     filters = listify(filters)
        #     dfs = [self.data[self.data['Station_Name'].str.contains(f, na='')] for f in filters]
        #
        #     df = pd.concat(dfs)
        # else:
        #     df = self.data
        df = self.data if qry_str is None else self.data.query(qry_str)

        funcs = listify(funcs)

        for idx, row in df.iterrows():
            if pd.isnull(row['Relative_Path']):
                print(f"Missing Relative Path for row: {row.name}")
                continue

            out_path = self.path/row['Relative_Path']
            if not out_path.exists():
                print(f'Folder {out_path} not found. Skipping')
                continue

            for func in funcs:
                func(row)

        print(f'Affected IDs: {df.index.to_list()}')

    def create_summary_radiometries(self, r_types=None, attrs=None):
        """Create the files xxx summary.csv for the r_types indicated. These files are necessary to load
        the mean radiometries of the database. The attributes will be stored in the final .csv
        this function.
        r_types are ['Rrs', 'Ed', 'Ld', 'Lu']
        """
        r_types = RadiometryDB.all_radiometries if r_types is None else listify(r_types)
        self.rdmtries = RadiometryDB._load_mean_radiometries(df=self.data,
                                                             db_anchor=self.path,
                                                             r_types=r_types,
                                                             path_column='Relative_Path',
                                                             attrs=attrs,
                                                             mean_fname=self.config['mean_interpolated']
                                                             )

        for r_type in r_types:
            fn = self.path / (r_type + self.config['summary'])
            print(f'Saving radiometry: {fn}')
            self.rdmtries[r_type].to_csv(fn, index=True)

        print(f'Radiometries {r_types} loaded in dictionary .rdmtries')

    def create_adjusted_reflectance(self, trios_main, qry_str=None):
        """
        Create the adjusted reflectance using Tristan's TRIOS software
        It is necessary to pass the main .py file from TRIOS with full path
        The code can be run for just a subset, using a query_string (qry_str)
        """
        def map_adjusted_reflectance(row):
            """Function to be applied to each row of the data table
            TRIOS from Tristan is executed as shell"""
            print(f"Processing: {row['Relative_Path']}")
            path = Path(self.path/row['Relative_Path'])
            os.system(f"echo '{path.as_posix()}'")
            args = "\"" + path.as_posix() + "\" 150 awr"
            args += f" --lat {row['Lat']} --lon {row['Long']}"
            args += " --odir \"" + path.as_posix() + "\""
            args += ' --format=csv --data_files="Ed_interpolated.csv Ld_interpolated.csv Lu_interpolated.csv"'
            args += " --ofile " + self.config['trios_output']
            args += " --plot --figdir \"" + path.as_posix() + "\""
            args += f" --altitude={row['Altitude']}"

            cmd = f"python \"{trios_main}\" {args}"
            os.system(f"echo {cmd}")
            os.system(f"python \"{trios_main}\" {args}")

            # After the execution of Trios, we open the trios csv and the mean_reflectances.csv and
            # append the results to the mean_reflectances.
            # open the resulting TRIOS csv file
            trios_df = pd.read_csv(path/self.config['trios_output'], header=[0, 1])

            # Grab just the reflectance and the date times
            trios_rrs = trios_df['Rrs'].drop(0)
            trios_dt = trios_df['param'].drop(0).rename(columns={'wl': 'DateTime'})
            trios_dt['DateTime'] = pd.to_datetime(trios_dt['DateTime'])

            # Save the adjusted Rrs (OSOAA) to the interpolated file
            rrs_osoaa = pd.concat([trios_dt, trios_rrs], axis=1)
            rrs_osoaa.to_csv(path / ('Rrs_OSOAA'+self.config['interpolated_name']), sep=';', index=False)

            # and calculate the mean
            trios_mean = trios_rrs.mean()
            trios_mean.name = 'Rrs_OSOAA'

            # open the Mean_Reflectance.csv, drop old Rrs_OSOAA and append the new
            df = pd.read_csv(path/self.config['mean_interpolated'], sep=';', index_col=0)
            df.drop(index='Rrs_OSOAA', errors='ignore', inplace=True)
            df = df.append(trios_mean)
            df['DateTime'] = df.loc['Rrs', 'DateTime']
            df.to_csv(path/self.config['mean_interpolated'], sep=';', index=True)

        # loop through the database and apply the mapping function
        self.loop_database(funcs=map_adjusted_reflectance, qry_str=qry_str)

    def reprocess_database(self, qry_str=None, prefixes=None, reflectance=False, mean_values=False, graphs=False):
        prefixes = prefixes if prefixes is not None else ['Rrs', 'Ed', 'Ld', 'Lu']

        def reprocess_db_callback(row):
            out_path = self.path / row['Relative_Path']

            rdmtry = Radiometry.from_folder(out_path, prefixes=prefixes, load_interpolated=True)

            if reflectance:
                rdmtry.create_reflectance()
                rdmtry.save_radiometries(save_backup=True, subset=False)

            if mean_values:
                rdmtry.save_mean_radiometry(subset=False)

            if graphs:
                rdmtry.save_radiometries_graph(rdmtry.folder, subset=False)

        self.loop_database(funcs=reprocess_db_callback, qry_str=qry_str)

    def update_metadata(self, attrs, qry_str=None, section='Metadata'):
        attrs = listify(attrs)

        def update_metadata_callback(row):
            out_path = self.path / row['Relative_Path']
            fn = out_path / 'metadata.txt'
            config = configparser.ConfigParser()
            config.read(fn)

            config[section].update({column: str(row[column]) for column in attrs})

            with open(fn, 'w') as configfile:
                config.write(configfile)

        self.loop_database(funcs=update_metadata_callback, qry_str=qry_str)

    def append_new_data(self, from_folder, copy_all_files=False, reprocess=False, overwrite=False,
                        base_name='spectrum_LO', sep='_', dir_level=0):
        """Append to the database new data that follows the structure root-Area-Station-Measurement
        The measurement folder will be renamed to match DB standard (yyyymmdd-hhmm)"""

        # Step 1 - Check the new data
        df = check_structure(from_folder, prefixes=['Ld', 'Ed', 'Lu'], base_name=base_name, sep=sep,
                             dir_level=dir_level)

        # Create the destination folder for each new measurement
        df['NewDir'] = self.path.as_posix() + '/' + df['Area'] + '/' + df['Station'] + '/' + df['Start_Date'].astype(
            'str').str.replace('-', '').str.replace(' ', '-').str[:14].str.replace(':', '')

        # Step 2 - Iterate through the rows of the dataframe to copy the old directory into the new directory
        pattern = '*' if copy_all_files else f'*{base_name}*'

        for idx, row in df.iterrows():
            copy_dir(row['Dir'], row['NewDir'], pattern=pattern)

        # Step 3 - If reprocess==True, reprocess radiometry to create Interpolated Reflectance, .PNG, etc
        if reprocess:
            for idx, row in df.iterrows():
                rdmtry = Radiometry.from_folder(row['NewDir'], prefixes=['Ed', 'Ld', 'Lu'])
                rdmtry.create_reflectance(min_wl=320, max_wl=950)
                rdmtry.save_radiometries(save_backup=False, subset=False)
                rdmtry.save_radiometries_graph(rdmtry.folder, subset=False)

                rdmtry.update_metadata()

        # Once everything is reprocessed, put the metadata into the resulting dataframe
        def map_get_attributes(folder):
            rdmtry2 = Radiometry.from_folder(folder)

            return pd.Series([rdmtry2.metadata['Metadata'].get('Metadata', 'ed_device'),
                              rdmtry2.metadata['Metadata'].get('Metadata', 'ld_device'),
                              rdmtry2.metadata['Metadata'].get('Metadata', 'lu_device'),
                              int(rdmtry2.metadata['Metadata'].get('Metadata', 'rrs_measurements'))])

        df['Ed_device'] = df['Ld_device'] = df['Lu_device'] = df['Rrs_measurements'] = None

        df[['Ed_device', 'Ld_device', 'Lu_device', 'Rrs_measurements']] = \
            pd.DataFrame(df['NewDir'].map(map_get_attributes).to_list())

        # step 4 - Create the columns to be updated in the control file
        df['Relative_Path'] = df['NewDir'].map(lambda x: Path(x).relative_to(self.path))
        df['Station_Name'] = df['Area'] + '-' + df['Station']
        df['Measurement'] = df['NewDir'].map(lambda x: Path(x).relative_to(self.path).name)

        # step 5 - Update the stations workbook
        for area, station in df.groupby(by=['Area', 'Station']).count().index:
            # self.create_station(area, station)
            self.insert_station(pd.Series({'Area': area,
                                           'Station': station,
                                           'Station_Name': '-'.join([area, station])}),
                                overwrite=overwrite)

        # step 6 - update the main workbook
        columns = ['Relative_Path', 'Station_Name', 'Measurement', 'Start_Date', 'End_Date', 'Ed_device',
                   'Ld_device', 'Lu_device', 'Rrs_measurements', 'Ed_measurements', 'Ld_measurements',
                   'Lu_measurements', 'Description']

        for _, row in df[columns].iterrows():
            self.insert_measurement(row, overwrite=overwrite)

        return 'OK'

    # Saving Functions
    def commit_workbook(self, workbook='main_workbook'):
        if workbook == 'stations_workbook':
            df = self.stations
        elif workbook == 'summertime_workbook':
            df = self.summertime
        else:
            df = self.data[['Relative_Path', 'Station_Name', 'Measurement', 'SPM', 'Chl-a', 'POC', 'DOC', 'Status',
                            'Start_Date','End_Date', 'gmt_datetime', 'Project', 'Ed_device', 'Ld_device',
                            'Lu_device', 'Rrs_measurements', 'Ed_measurements', 'Ld_measurements',
                            'Lu_measurements', 'Description']]

        append_df_to_excel(self.control_file, df, sheet_name=self.config[workbook], startrow=0,
                           truncate_sheet=True)

    def commit_all(self, workbooks=None):
        """Commit changes to the control file. Workbooks argument specifies which workbook to commit
        Options are: 'main_workbook', 'stations_workbook', 'summertime_workbook'"""

        workbooks = ['main_workbook', 'stations_workbook', 'summertime_workbook'] if workbooks is None \
            else listify(workbooks)

        for workbook in workbooks:
            self.commit_workbook(workbook=workbook)

    # Adjust GMT Times
    @staticmethod
    def in_summer_period(dt, summer_times):
        def in_period(row):
            return (dt > row['StartTime']) and (dt < row['EndTime'])

        return summer_times.apply(in_period, axis=1).any()

    @staticmethod
    def map_gmt_time(row, date_column, gmt_column, use_st_column, summer_times):
        dt = row[date_column]
        is_summer = RadiometryDB.in_summer_period(dt, summer_times=summer_times)

        apply_summertime = (row[use_st_column] == 1) and is_summer

        delta = -row[gmt_column] if not apply_summertime else -(row[gmt_column] + 1)

        return dt + pd.to_timedelta(delta, unit='h')

    def fill_gmt_datetime(self):
        map_func = partial(RadiometryDB.map_gmt_time, date_column='Start_Date', gmt_column='GMT',
                           use_st_column='SummerTime', summer_times=self.summertime)

        self.data['gmt_datetime'] = self.data.apply(map_func, axis=1)

    def __len__(self):
        return len(self.data) if self.data is not None else 0

    def __repr__(self):
        s = f'Radiometry Database with {len(self)} records and ' \
            f'{len(self.stations) if self.stations is not None else 0} stations'
        return s

